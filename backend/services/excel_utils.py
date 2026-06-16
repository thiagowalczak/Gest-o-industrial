"""Funções utilitárias para leitura de planilhas Excel enviadas pelos usuários."""
from __future__ import annotations
import calendar
import io
import unicodedata
from datetime import datetime
from typing import Optional

import openpyxl

_MESES = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}

_LINHAS_IGNORADAS = {
    "saldo inicial", "entradas (receitas)", "saidas (despesas)",
    "total receitas", "total saidas", "resultado do mes",
    "total entradas", "total despesas",
}

_LINHAS_ENTRADAS = {"entradas (receitas)", "total receitas"}
_LINHAS_SAIDAS = {"saidas (despesas)", "total saidas"}


def is_fluxo_caixa(conteudo: bytes) -> bool:
    """Retorna True se a planilha parece ser um Fluxo de Caixa (primeira célula contém 'fluxo de caixa')."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active
    primeira = normalizar(ws.cell(1, 1).value or "")
    return "fluxo de caixa" in primeira


def ler_fluxo_caixa(conteudo: bytes) -> list[dict]:
    """
    Lê planilha no formato Fluxo de Caixa (categorias x meses) e retorna
    uma lista de lançamentos individuais, um por (categoria, mês) com valor != 0.
    Cada item inclui a chave 'tipo': 'receber' (ENTRADAS) ou 'pagar' (SAÍDAS).
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active

    # Extrai ano do cabeçalho (ex: "FLUXO DE CAIXA - 2026")
    cabecalho = str(ws.cell(1, 1).value or "")
    ano = datetime.now().year
    for parte in cabecalho.split():
        if parte.isdigit() and len(parte) == 4:
            ano = int(parte)
            break

    # Mapeia índice de coluna → número do mês
    col_mes: dict[int, int] = {}
    for col_idx, cel in enumerate(ws[1], start=1):
        val = str(cel.value or "").strip().upper()
        if val in _MESES:
            col_mes[col_idx] = _MESES[val]

    if not col_mes:
        return []

    tipo_atual = "receber"
    lancamentos = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        categoria = normalizar(row[0])
        if not categoria:
            continue
        if categoria in _LINHAS_SAIDAS:
            tipo_atual = "pagar"
            continue
        if categoria in _LINHAS_ENTRADAS:
            tipo_atual = "receber"
            continue
        if categoria in _LINHAS_IGNORADAS:
            continue

        nome_categoria = str(row[0]).strip()

        for col_idx, mes in col_mes.items():
            valor = row[col_idx - 1]
            try:
                valor = float(valor or 0)
            except (TypeError, ValueError):
                continue
            if valor == 0:
                continue

            ultimo_dia = calendar.monthrange(ano, mes)[1]
            vencimento = f"{ano}{mes:02d}{ultimo_dia:02d}"
            emissao = f"{ano}{mes:02d}01"
            mes_nome = list(_MESES.keys())[mes - 1]
            titulo = f"{nome_categoria[:18]}-{mes_nome}{ano}"  # max 30 chars

            lancamentos.append({
                "titulo": titulo,
                "contraparte_nome": nome_categoria,
                "emissao": emissao,
                "vencimento": vencimento,
                "valor": abs(valor),
                "saldo": abs(valor),
                "tipo_doc": "FC",  # String(10) no banco
                "tipo": tipo_atual,
            })

    return lancamentos


def normalizar(texto) -> str:
    texto = str(texto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return texto


def converter_data(valor) -> str:
    """Converte célula de data para o formato AAAAMMDD usado internamente."""
    if isinstance(valor, datetime):
        return valor.strftime("%Y%m%d")
    texto = str(valor or "").strip()
    if not texto:
        return ""
    if "/" in texto:
        partes = texto.split("/")
        if len(partes) == 3:
            d, m, a = partes
            if len(a) == 2:
                a = "20" + a
            return f"{a}{m.zfill(2)}{d.zfill(2)}"
    if "-" in texto and len(texto) == 10:
        a, m, d = texto.split("-")
        return f"{a}{m.zfill(2)}{d.zfill(2)}"
    return texto


def formatar_data_br(aaaammdd) -> str:
    """Converte uma data interna (AAAAMMDD) para exibição em DD/MM/AAAA."""
    texto = str(aaaammdd or "").strip()
    if len(texto) == 8 and texto.isdigit():
        return f"{texto[6:8]}/{texto[4:6]}/{texto[0:4]}"
    return ""


def gerar_csv(colunas: list[str], linhas: list[list]) -> bytes:
    """
    Gera um arquivo .csv (em memória), com separador ';' e BOM UTF-8,
    para abrir corretamente em Excel/Excel para a Web no padrão brasileiro.
    """
    import csv

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(colunas)
    writer.writerows(linhas)
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def gerar_modelo_excel(colunas: list[str], exemplos: list[list]) -> bytes:
    """
    Gera um arquivo .xlsx (em memória) com uma linha de cabeçalho
    formatada e uma ou mais linhas de exemplo, para servir como
    modelo de importação para os usuários.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"

    cabecalho_fill = PatternFill(start_color="FF7A1E", end_color="FF7A1E", fill_type="solid")
    cabecalho_font = Font(bold=True, color="FFFFFF")

    for col_idx, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx, value=titulo)
        celula.font = cabecalho_font
        celula.fill = cabecalho_fill
        celula.alignment = Alignment(horizontal="center")
        ws.column_dimensions[celula.column_letter].width = max(18, len(titulo) + 2)

    for row_idx, linha in enumerate(exemplos, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def ler_linhas(conteudo: bytes, mapa_colunas: dict, planilha: Optional[str] = None, nome_arquivo: Optional[str] = None) -> list[dict]:
    """
    Lê um arquivo .xlsx/.xls ou .csv e retorna uma lista de dicionários,
    mapeando os cabeçalhos da planilha (normalizados) para as chaves
    internas definidas em `mapa_colunas` (ex.: {"codigo do produto": "codigo", ...}).
    Colunas não reconhecidas são ignoradas.
    """
    if nome_arquivo and nome_arquivo.lower().endswith(".csv"):
        return _ler_linhas_csv(conteudo, mapa_colunas)

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb[planilha] if planilha and planilha in wb.sheetnames else wb.active

    chaves_colunas = [mapa_colunas.get(normalizar(cel.value)) for cel in ws[1]]

    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        item = {}
        for chave, valor in zip(chaves_colunas, row):
            if chave:
                item[chave] = valor
        if item:
            linhas.append(item)
    return linhas


def _ler_linhas_csv(conteudo: bytes, mapa_colunas: dict) -> list[dict]:
    import csv

    texto = conteudo.decode("utf-8-sig")
    try:
        delimitador = csv.Sniffer().sniff(texto[:2048], delimiters=";,").delimiter
    except csv.Error:
        delimitador = ";"

    linhas_brutas = list(csv.reader(io.StringIO(texto), delimiter=delimitador))
    if not linhas_brutas:
        return []

    chaves_colunas = [mapa_colunas.get(normalizar(cel)) for cel in linhas_brutas[0]]

    linhas = []
    for row in linhas_brutas[1:]:
        if not any(str(v).strip() != "" for v in row):
            continue
        item = {}
        for chave, valor in zip(chaves_colunas, row):
            if chave:
                item[chave] = valor
        if item:
            linhas.append(item)
    return linhas
